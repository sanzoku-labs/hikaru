from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

from app.models.schemas import ColumnInfo, DataSchema


class DataProcessor:
    @staticmethod
    def parse_file(
        file_path: str, file_extension: str, sheet_name: Optional[str | int] = None
    ) -> pd.DataFrame:
        """Parse CSV or Excel file into DataFrame.

        Supports both US and European formats.

        Args:
            file_path: Path to the file
            file_extension: File extension (csv, xlsx, xls)
            sheet_name: For Excel files, the sheet name or index
                (0-based). Defaults to 0 (first sheet)
        """
        if file_extension == "csv":
            # Try US format first (comma delimiter, period decimal)
            try:
                df = pd.read_csv(file_path, encoding="utf-8")

                # Check if parsing was successful (more than 1 column detected)
                if len(df.columns) > 1:
                    return df

                # If only 1 column, likely European format with semicolons
                raise ValueError("Single column detected, trying European format")

            except (ValueError, pd.errors.ParserError):
                # Try European format (semicolon delimiter, comma decimal)
                try:
                    df = pd.read_csv(
                        file_path,
                        encoding="utf-8",
                        sep=";",
                        decimal=",",
                        thousands=None,  # Disable thousands separator
                    )
                    return df
                except Exception as e:
                    # If both fail, raise the original error
                    raise ValueError(
                        "Failed to parse CSV file. Tried both US (comma) "
                        f"and European (semicolon) formats. Error: {e}"
                    )

        elif file_extension in ["xlsx", "xls"]:
            # Default to first sheet if not specified
            sheet = sheet_name if sheet_name is not None else 0
            return pd.read_excel(file_path, sheet_name=sheet)
        else:
            raise ValueError(f"Unsupported file extension: {file_extension}")

    @staticmethod
    def get_excel_sheets(file_path: str) -> List[Dict[str, Any]]:
        """Get list of all sheets in an Excel file with metadata.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of dicts containing sheet metadata: name, index, is_hidden, row_count, column_count
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel sheet detection. Install with: pip install openpyxl"
            )

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = []

            for idx, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                sheets.append(
                    {
                        "name": sheet_name,
                        "index": idx,
                        "is_hidden": sheet.sheet_state == "hidden",
                        "row_count": sheet.max_row,
                        "column_count": sheet.max_column,
                    }
                )

            workbook.close()
            return sheets
        except Exception as e:
            raise ValueError(f"Failed to read Excel file sheets: {str(e)}")

    @staticmethod
    def validate_dataframe(df: pd.DataFrame) -> Tuple[bool, Optional[str]]:
        """Validate dataframe meets requirements"""
        # Row count check
        if len(df) > 100_000:
            return False, f"Too many rows: {len(df)} (limit: 100,000)"

        if len(df) < 2:
            return False, "Insufficient data (minimum: 2 rows)"

        # Column count check
        if len(df.columns) > 50:
            return False, f"Too many columns: {len(df.columns)} (limit: 50)"

        # Check for at least one numeric column
        numeric_cols = df.select_dtypes(include=["number"]).columns
        if len(numeric_cols) == 0:
            return False, "No numeric columns found for visualization"

        return True, None

    @staticmethod
    def detect_column_type(series: pd.Series) -> str:
        """Detect column type: numeric, categorical, or datetime"""
        # Check for datetime
        if pd.api.types.is_datetime64_any_dtype(series):
            return "datetime"

        # Check for time-related column names (even if numeric)
        # This catches month_id (202505), week_id (202518), etc.
        if series.name:
            col_name_lower = str(series.name).lower()
            time_keywords = ["month", "week", "year", "quarter", "day", "date", "period"]
            if any(keyword in col_name_lower for keyword in time_keywords):
                # Likely a time dimension (even if stored as integer)
                return "datetime"

        # Try to convert to datetime
        if series.dtype == "object":
            try:
                pd.to_datetime(series.dropna().head(100), format="mixed")
                return "datetime"
            except (ValueError, TypeError):
                pass

        # Check for numeric
        if pd.api.types.is_numeric_dtype(series):
            return "numeric"

        # Default to categorical
        return "categorical"

    @staticmethod
    def _sanitize_value(value: Any) -> Any:
        """Convert NaN/Inf values to None for JSON serialization"""
        # Check for NaN first (works on all types)
        if pd.isna(value):
            return None

        # Only check for infinity on numeric types
        if isinstance(value, (int, float, np.integer, np.floating)):
            if np.isinf(value):
                return None

        return value

    @staticmethod
    def analyze_schema(df: pd.DataFrame) -> DataSchema:
        """Analyze dataframe and generate schema"""
        columns_info = []

        for col in df.columns:
            series = df[col]
            col_type = DataProcessor.detect_column_type(series)

            # Sanitize sample values (remove NaN)
            sample_values = [
                DataProcessor._sanitize_value(v) for v in series.dropna().head(5).tolist()
            ]

            col_info = ColumnInfo(
                name=col,
                type=col_type,
                null_count=int(series.isnull().sum()),
                unique_values=int(series.nunique()) if col_type == "categorical" else None,
                sample_values=sample_values,
            )

            # Add numeric stats (with NaN handling)
            if col_type == "numeric":
                col_info.min = DataProcessor._sanitize_value(float(series.min()))
                col_info.max = DataProcessor._sanitize_value(float(series.max()))
                col_info.mean = DataProcessor._sanitize_value(float(series.mean()))
                col_info.median = DataProcessor._sanitize_value(float(series.median()))

            columns_info.append(col_info)

        # Get preview (first 10 rows) - replace NaN/Inf with None
        preview = df.head(10).replace([np.nan, np.inf, -np.inf], None).to_dict("records")

        return DataSchema(columns=columns_info, row_count=len(df), preview=preview)
